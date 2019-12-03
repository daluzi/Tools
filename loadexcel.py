# _*_ coding: utf-8 _*_
# @Author   : daluzi
# @time     : 2019/12/2 22:53
# @File     : loadexcel.py
# @Software : PyCharm

import numpy as np
import xlrd
from openpyxl import load_workbook

def loadExcel_1(filename):
	workbook = xlrd.open_workbook(filename)
	print(workbook.sheet_names())  # 查看所有sheet
	booksheet = workbook.sheet_by_index(0)  # 用索引取第一个sheet
	booksheet = workbook.sheet_by_name('Sheet1')  # 或用名称取sheet
	# 读单元格数据
	cell_11 = booksheet.cell_value(0, 0)
	cell_21 = booksheet.cell_value(1, 0)
	# 读一行数据
	row_3 = booksheet.row_values(2)
	print(np.array(row_3))
	print(row_3[6])

def loadExcel_2(filename):
	workbook = load_workbook(filename)
	# booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
	sheets = workbook.get_sheet_names()  # 从名称获取sheet
	booksheet = workbook.get_sheet_by_name(sheets[0])

	rows = booksheet.rows
	columns = booksheet.columns
	KDATop = []
	KDASingle = []
	KDAAdc = []
	KDAJungle = []
	KDASupport = []
	# 迭代所有的行
	for row in rows:
		line = [col.value for col in row]
		if line[5] == "上单":
			KDATop.append(line[3])
			continue
		elif line[5] == "中单":
			KDASingle.append(line[3])
			continue
		elif line[5] == "ADC":
			KDAAdc.append(line[3])
			continue
		elif line[5] == "打野":
			KDAJungle.append(line[3])
			continue
		elif line[5] == "辅助":
			KDASupport.append(line[3])
			continue
	print("Top:\n",KDATop)
	print("Single:\n",KDASingle)
	print("Adc:\n",KDAAdc)
	print("Jungle:\n",KDAJungle)
	print("support:\n",KDASupport)
	# # 通过坐标读取值
	# cell_11 = booksheet.cell('A1').value
	# cell_11 = booksheet.cell(row=1, column=1).value

loadExcel_2('player.xlsx')